import os
import time

import openpyxl
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter

from config.filePath import filePath
from lib.yamlMethod import readYaml


def picfield():
    for i in os.listdir(filePath.FPICTURES_DIR):
        yield i

def readtxt():
    with open(filePath.DESCRIPTION, 'r', encoding='gb18030') as f:
        for i in f.readlines():
            yield i
pic = picfield()
def writeExcel(sport):
    '''
    将用例结果写到excel
    '''
    description = readtxt()
    print(description)
    excelname = 'hhh.xlsx'
    wb = openpyxl.Workbook()
    wb.save(excelname)
    wb = load_workbook('hhh.xlsx')
    sheet = wb.worksheets[0]
    nowrow = sheet.max_row + 1
    print('当前行', sheet.max_row)
    sheet.cell(nowrow, 1).value = sport
    piclen = readYaml('sport.yaml')[sport]
    sheet.row_dimensions[nowrow].height = 400
    for i in range(2, len(piclen) + 2):
        sheet[get_column_letter(i) + str(nowrow)] = next(description)
        if piclen[i - 2] == 's':
            continue
        apic = os.path.join(filePath.FPICTURES_DIR, next(pic))
        if piclen[i - 2] == 'a':
            img = Image.open(apic)
            out = img.resize((int((img.size[0] * 0.35)), int((img.size[1]) * 0.35)))
            out.save(apic)
        sheet.column_dimensions[get_column_letter(i)].width = 40
        img = image.Image(apic)
        sheet.add_image(img, get_column_letter(i) + str(nowrow))
    wb.save('hhh.xlsx')
    wb.close()
if __name__ == '__main__':
    writeExcel('登山')