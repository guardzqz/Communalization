import os
import time

import openpyxl
import uiautomator2 as u2
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter
from ble.menst_cmd import MenstCmd
from config.filePath import filePath
from lib.OtherMethod import writeType
from lib.getConfig import getConfig
from lib.yamlMethod import writeYaml, readYaml
from method.OperationDir import clearStatus, delDir
from photo.mainPhoto import mainPhoto

class gt02_menst:
    def __init__(self):
        pass
        self.demo = u2.connect(getConfig('deviceid', 'app'))
        self.device = mainPhoto()
        self.PJ_name = 'gt02'
        self.yamlpath = 'menst.yaml'
        clearStatus()
        writeYaml(self.yamlpath)
        self.excelname = 'gt02_menst'
        self.newExcel()
        self.watch_cmd = readYaml(filePath.watch_cmd)[self.PJ_name]
        self.slide_up = self.watch_cmd['上滑']
        self.slide_down = self.watch_cmd['下滑']
        self.slide_left = self.watch_cmd['左滑']
        self.slide_right = self.watch_cmd['右滑']

    def demotext(self, text):
        time.sleep(0.5)
        self.demo.app_start(getConfig('demoname', 'app'))
        time.sleep(0.5)
        if self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').exists(1):
            self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').click()
        self.demo(resourceId = 'test.com.ido:id/custom_cmd_data_et').clear_text()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').set_text(text)
        self.demo.press('back')
        self.demo(text='发送自定义命令').click()

    def readtxt(self):
        with open(filePath.DESCRIPTION, 'r', encoding='utf-8') as f:
            for i in f.readlines():
                yield i

    def initial_menst(self,case_NO, initialNO ='initial'):
        menst_yaml = readYaml(filePath.menstPara)[case_NO][initialNO]
        off_on = menst_yaml['off_on']
        menstrual_lenth = menst_yaml['menstrual_lenth']
        menstrual_cycle = menst_yaml['menstrual_cycle']
        date = menst_yaml['date']
        ovulation_interval_day = menst_yaml['ovulation_interval_day']
        ovulation_before_day = menst_yaml['ovulation_before_day']
        ovulation_after_day = menst_yaml['ovulation_after_day']
        notify_flag = menst_yaml['notify_flag']
        init_cmd = MenstCmd.menses_set_cmd(off_on, menstrual_lenth, menstrual_cycle, date, ovulation_interval_day, ovulation_before_day, ovulation_after_day, notify_flag)
        set_date_cmd = MenstCmd.set_time(menst_yaml['dateing'])
        self.demotext(init_cmd)
        self.demotext(set_date_cmd)

    def process_cmd(self,mode):
        process = readYaml(filePath.watch_cmd)[self.PJ_name][mode]
        for cmd in process:
            if type(cmd) == str:
                self.demotext(cmd)
            elif type(cmd) == int:
                time.sleep(cmd)

    def menst_case1(self):
        self.initial_menst('case_data_1')
        self.initial_menst('case_data_1')
        date_list = readYaml(filePath.menstPara)['case_data_1']['date']
        entry_list = readYaml(filePath.menstPara)['case_data_1']['entry']
        for j in range(len(date_list)):
            self.demotext(MenstCmd.set_time(date_list[j]))
            time.sleep(6)
            self.device.camera(self.yamlpath, date_list[j]+entry_list[j])

    def menst_case2(self):
        self.initial_menst('case_data_2')
        self.initial_menst('case_data_2')
        date_list = readYaml(filePath.menstPara)['case_data_2']['date']
        entry_list = readYaml(filePath.menstPara)['case_data_2']['entry']
        his_list = readYaml(filePath.menstPara)['case_data_2']['his']
        his_one_cmd = MenstCmd.menses_his_cmd([readYaml(filePath.menstPara)['case_data_2']['his'][0]])
        self.demotext(his_one_cmd)
        self.demotext(self.slide_up) #上滑
        time.sleep(1)
        self.device.camera(self.yamlpath, entry_list[0])
        time.sleep(1)
        self.demotext(self.slide_up) #上滑
        time.sleep(1)
        self.device.camera(self.yamlpath, entry_list[1])
        time.sleep(1)
        his_cmd = MenstCmd.menses_his_cmd(his_list)
        self.demotext(his_cmd)
        self.demotext(self.slide_down) #下滑
        time.sleep(1)
        self.device.camera(self.yamlpath, entry_list[2])
        time.sleep(1)
        self.demotext(self.slide_up)  # 上滑
        time.sleep(1)
        self.device.camera(self.yamlpath, entry_list[3])
        self.demotext(self.slide_down)  # 下滑
        self.demotext(self.slide_down)  # 下滑
        self.demotext(self.slide_down)  # 下滑
        self.demotext(self.slide_down)  # 下滑


    def menst_case3(self):
        pass

    def stop_go_menst_ui(self):
        self.process_cmd('end_stop')
        self.process_cmd('go_menst')
        pass

    def menst_caseNO1(self, case_data_NO):
        #主提醒
        self.initial_menst(case_data_NO)
        self.initial_menst(case_data_NO)
        case_data = readYaml(filePath.menstPara)[case_data_NO]
        date_list = case_data['date']
        entry_list = case_data['entry']
        entry_list2 = case_data['entry2']
        trnslate_list = case_data['translate']
        rem_time = case_data['rem_time']
        rem_cmd = MenstCmd.menses_remind_cmd(rem_time['hour'], rem_time['start_day'], rem_time['ovulation_day'], rem_time['pregnancy_day_before'], rem_time['pregnancy_day_end'], rem_time['menstrual_day_end'])
        self.demotext(rem_cmd)
        for i in range(len(date_list)):
            self.demotext(MenstCmd.set_time(date_list[i]))
            time.sleep(3)
            print(entry_list2[i])
            self.device.camera(self.yamlpath, date_list[i]+entry_list2[i])
            time.sleep(2)
            self.device.camera(self.yamlpath, date_list[i]+entry_list[i])
            time.sleep(1)
            if trnslate_list[i]:
                self.demotext(self.slide_right)

        # #跨月提醒
        self.initial_menst(case_data_NO,'initial2')
        case_data = readYaml(filePath.menstPara)[case_data_NO]
        date2_list = case_data['date2']
        entry3_list = case_data['entry3']
        rem2_time = case_data['rem_time2']
        rem2_cmd = MenstCmd.menses_remind_cmd(rem2_time['hour'], rem2_time['start_day'], rem2_time['ovulation_day'], rem2_time['pregnancy_day_before'], rem2_time['pregnancy_day_end'], rem2_time['menstrual_day_end'])
        self.demotext(rem2_cmd)
        for i in range(2):
            self.demotext(MenstCmd.set_time(date2_list[i]))
            time.sleep(5)
            self.device.camera(self.yamlpath, entry3_list[i])
            time.sleep(2)
            self.demotext(self.slide_right)
        if case_data_NO != 'case_data_67' and case_data_NO != 'case_data_68':
        #消息中提醒
            self.demotext('74-0b-01')
            self.demotext('74-07-ff-ff')
            self.demotext(self.slide_right)
            self.demotext(self.slide_right)
            self.demotext(MenstCmd.set_time(date2_list[2]))
            time.sleep(5)
            self.demotext(self.slide_right)
            self.demotext(self.slide_down)
            time.sleep(1)
            self.device.camera(self.yamlpath, entry3_list[2])
            # self.demotext('74-01-00-f0-00-f0-00-06')

            #运动中提醒
            self.demotext('74-07-ff-ff')
            self.process_cmd('start_sotp')
            self.demotext(MenstCmd.set_time(date2_list[3]))
            time.sleep(6)
            self.device.camera(self.yamlpath, entry3_list[3])
            time.sleep(1)
            self.demotext(self.slide_right)
            self.stop_go_menst_ui()

    def test_menst(self):
        self.demotext("74-0b-01")
        self.demotext("74-07-ff-ff")
        case_list = ['case_data_32', 'case_data_42', 'case_data_50', 'case_data_58', 'case_data_66', 'case_data_67', 'case_data_68']
        # case_fun_list = [self.menst_case1, self.menst_case2]
        case_fun_list = []
        fun_n = len(case_fun_list)
        case_all = case_fun_list+case_list
        case_txt_list = ['经期开始提醒功能检查', '经期结束提醒功能检查', '易孕期开始提醒功能检查', '易孕期结束提醒功能检查', '排卵日提醒功能检查1', '排卵日提醒功能检查2',
                         '排卵日提醒功能检查3','经期跟踪各阶段文案检查','经期周期跟踪图表检查']
        for i in range(len(case_all)):
            pic = self.picfield()
            clearStatus()
            # self.device.recodingScreen()
            data = readYaml(self.yamlpath)
            if data is None:
                data = {}
            data[case_txt_list[i]] = []
            writeYaml(self.yamlpath, data)
            # 执行用例
            if i < fun_n:
                case_all[i]()
            else:
                self.menst_caseNO1(case_all[i])
            time.sleep(3)
            self.writeExcel(pic, case_txt_list[i])

    def newExcel(self):
        self.excelname = self.excelname + str(time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())) + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(self.excelname)

    def writeExcel(self, pic, sport):
        '''
        将用例结果写到excel
        '''
        description = self.readtxt()
        wb = load_workbook(self.excelname)
        sheet = wb.worksheets[0]
        nowrow = sheet.max_row+1
        print('当前行', sheet.max_row)
        sheet.cell(nowrow, 1).value = sport
        piclen = readYaml(self.yamlpath)[sport]
        sheet.row_dimensions[nowrow].height = 250
        for i in range(2, len(piclen) + 2):
            sheet[get_column_letter(i) + str(nowrow)] = next(description)
            if piclen[i-2] == 's':
                continue
            apic = os.path.join(filePath.FPICTURES_DIR, next(pic))
            if piclen[i - 2] == 'a':
                img = Image.open(apic)
                out = img.resize((int((img.size[0] * 0.35)), int((img.size[1]) * 0.35)))
                out.save(apic)
            sheet.column_dimensions[get_column_letter(i)].width = 40
            img = image.Image(apic)
            sheet.add_image(img, get_column_letter(i) + str(nowrow))
        wb.save(self.excelname)
        wb.close()

    def picfield(self):
        for i in os.listdir(filePath.FPICTURES_DIR):
            yield i

if __name__ == '__main__':
    gt02_menst().test_menst()
