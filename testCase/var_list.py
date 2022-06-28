import os
import time
import openpyxl
import ruamel
import re
import uiautomator2 as u2
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter
from ble.eventremin_cmd import eventCmd
from config.filePath import filePath
from lib.OtherMethod import writeType
from lib.getConfig import getConfig
from lib.yamlMethod import writeYaml, readYaml
from method.OperationDir import clearStatus, delDir
from photo.mainPhoto import mainPhoto

class Basic_modules:

    def __init__(self):
        self.demo = u2.connect(getConfig('deviceid', 'app'))
        self.device = mainPhoto()
        self.PJ_name = 'IDW_03'   #项目名称
        self.yamlpath = 'var_list.yaml'
        self.yamlNO = 'test_NO'
        self.test_data_NO_list = []
        self.data = {}
        clearStatus()
        writeYaml(self.yamlpath)
        self.excelname = self.PJ_name
        self.newExcel()
        self.watch_cmd = readYaml(filePath.var_list)[self.PJ_name]
        self.h_unit = '12'   #时制
        self.demo_apkname = getConfig('demoname', 'app')
        pass


    def demotext(self, text):
        time.sleep(0.5)
        self.demo.app_start(getConfig('demoname', 'app'))
        time.sleep(0.5)
        if self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').exists(1):
            self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').click()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').clear_text()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').set_text(text)
        self.demo.press('back')
        self.demo(text='发送自定义命令').click()

    def readtxt(self):
        with open(filePath.DESCRIPTION, 'r', encoding='utf-8') as f:
            for i in f.readlines():
                yield i
                
    def ten_to_hex(self, num, several=0):
        """
        num为待转换为16进制数据，
        text为excel单元格中读取到的命令
        serveral 为uint32_t value
        """
        if several == 0:
            several = len(str(num))
            if len(str(num)) % 2 == 1:
                num = "0" + str(num)
                several = len(str(num))
        trans_code = '{:0>{several}x}'.format(int(num), several=several)  # 小端序  several设置位数
        instruct = ''
        for i in range(len(trans_code), 0, -2):
            instruct = "-".join([instruct, trans_code[i - 2:i]])
        return instruct[1:]


    def new_excel_row(self):
        self.pic = self.picfield()
        clearStatus()
        self.data = readYaml(self.yamlpath) if readYaml(self.yamlpath) else {}
        self.lang = self.language.pop(0)   # 当前语言
        self.data[self.lang] = []
        writeYaml(self.yamlpath, self.data)

    def save_excel_row(self):
        self.writeExcel(self.pic, self.lang)

    def newExcel(self):
        self.excelname = self.excelname + str(time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())) + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(self.excelname)

    def writeExcel(self, pic, sport):
        '''
        将用例结果写到excel
        '''
        description = self.readtxt()
        wb = load_workbook(self.excelname)
        sheet = wb.worksheets[0]
        nowrow = sheet.max_row + 1
        print('当前行', sheet.max_row)
        sheet.cell(nowrow, 1).value = sport
        piclen = readYaml(self.yamlpath)[sport]
        sheet.row_dimensions[nowrow].height = 250
        for i in range(2, len(piclen) + 2):
            sheet[get_column_letter(i) + str(nowrow)] = next(description)
            if piclen[i - 2] == 's':
                continue
            apic = os.path.join(filePath.FPICTURES_DIR, next(pic))
            if piclen[i - 2] == 'a':
                img = Image.open(apic)
                out = img.resize((int((img.size[0] * 0.35)), int((img.size[1]) * 0.35)))
                out.save(apic)
            sheet.column_dimensions[get_column_letter(i)].width = 40
            img = image.Image(apic)
            sheet.add_image(img, get_column_letter(i) + str(nowrow))
        wb.save(self.excelname)
        wb.close()

    def picfield(self):
        for i in os.listdir(filePath.FPICTURES_DIR):
            yield i


#各类执行方法

    '''上滑'''
    def slide_up(self):
        self.demotext(self.watch_cmd['上滑'])

    '''下滑'''
    def slide_down(self):
        self.demotext(self.watch_cmd['下滑'])

    '''左滑'''
    def slide_left(self):
        self.demotext(self.watch_cmd['左滑'])

    '''右滑'''
    def slide_right(self):
        self.demotext(self.watch_cmd['右滑'])

    '''右键常规'''
    def ringt_click_1(self):
        self.demotext(self.watch_cmd['右键'])

    '''右键2次唤醒'''
    def ringt_click_2(self):
        self.demotext(self.watch_cmd['右键0'])
        self.demotext(self.watch_cmd['右键1'])
    '''右键长按'''
    def ringt_press(self):
        self.demotext(self.watch_cmd['长按右键'])

    '''点击任意坐标'''
    def click(self,x_y):
        x, y = hex(int(x_y[0]))[2:], hex(int(x_y[1]))[2:]
        x = '0' * (4 - len(x)) + x
        y = '0' * (4 - len(y)) + y
        s = '7401%s%s0001'% (x, y)
        appcmd="-".join(re.findall(r".{2}", s))
        self.demotext(appcmd)

    '''任意界面返回主页'''
    def return_home(self,i):
        for _ in range(int(i)):
            self.demotext(self.watch_cmd['右滑'])


    '''返回对应指定键值'''
    def get_key_value(self, yamldict):
        for key in yamldict.keys():
            return key, yamldict[key]


    '''拍照延时'''
    def sleep_camera(self, yamlpath, text, s1=1, s2=1):
        time.sleep(s1)
        self.device.camera(yamlpath, text)
        time.sleep(s2)

    '''任意延时'''
    def sleep(self):
        time.sleep(1)

    '''设置手表语言'''
    def set_language(self,language):
        dict = {'英文':'02','德文':'04','西班牙文':'06','法文':'03','意大利文':'05','葡萄牙文':'17'}
        frist = '03-11-00-00-00-00-'
        results = frist + dict[language]
        self.demotext(results)

    '''设置手表单位'''
    def set_unit(self,unit):
        if unit == '公制':
            self.demotext(self.watch_cmd['公制'])
        elif unit == '英制':
            self.demotext(self.watch_cmd['英制'])

    '''设置手表时间'''
    def set_time(self,date):
        '''
        设置时间,格式1: 2021/12/27/09/00'
                格式2: 2021/12/27
        '''
        if len(date) > 11:
            years, months, days, hour, minute,= date.split('/')
            cmd = "03-01-" + self.ten_to_hex(years) + "-" + self.ten_to_hex(months) + "-" + self.ten_to_hex(days) + "-" + self.ten_to_hex(hour) + "-" + self.ten_to_hex(minute) + "-38-ff"
        else:
            years, months, days, = date.split('/')
            cmd = "03-01-" + self.ten_to_hex(years) + "-" + self.ten_to_hex(months) + "-" + self.ten_to_hex(days) + "-09-00-38-ff"
        self.demotext(cmd)



    '''设置时制'''
    def set_h_unit(self,h_unit):
        if h_unit in ('12',12 ):
            self.demotext('03-11-00-00-00-1e-00-02-1e-00-01')
        elif h_unit in ('24',24 ):
            self.demotext('03-11-00-00-00-1e-00-01-1e-00-01')


    '''翻转时制'''
    def reverse_h_unit(self):
        if self.h_unit == '12':
            self.h_unit = '24'
            self.set_h_unit(self.h_unit)
        else:
            self.h_unit = '12'
            self.set_h_unit(self.h_unit)











