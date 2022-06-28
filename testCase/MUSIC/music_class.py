import os
import time
import openpyxl
import ruamel
import re
import uiautomator2 as u2
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter
from ble.eventremin_cmd import eventCmd
from config.filePath import filePath
from lib.OtherMethod import writeType
from lib.getConfig import getConfig
from lib.yamlMethod import writeYaml, readYaml
from method.OperationDir import clearStatus, delDir
from photo.mainPhoto import mainPhoto

'''
    说明：
        1. 编码：longt
        2. 开始时间：2022-6-13
        3. 最近更新：2020-6-13
        4. 简介：音乐模块基本类
'''

class Music:


    def __init__(self):


        self.demo = u2.connect(getConfig('deviceid4', 'app'))
        self.device = mainPhoto()
        self.PJ_name = 'ID_W03'   # 项目名称
        self.music = 'music'    # 模块名称
        self.yamlpath = 'music_carema.yaml'
        self.yamlNO = 'music_NO'
        self.music_data_NO_list = []
        self.data = {}
        clearStatus()
        writeYaml(self.yamlpath)
        self.excelname = self.PJ_name+'_music'
        self.newExcel()
        self.watch_cmd = readYaml(filePath.music_cmd)[self.music]
        self.h_unit = '12'   #时制
        self.demo_apkname = getConfig('demoname', 'app')
        pass


    def demotext(self, text):

        time.sleep(0.5)
        self.demo.app_start(getConfig('demoname', 'app'))
        time.sleep(0.5)
        if self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').exists(1):
            self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').click()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').clear_text()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').set_text(text)
        self.demo.press('back')
        self.demo(text='发送自定义命令').click()


    def readtxt(self):
        with open(filePath.DESCRIPTION, 'r', encoding='utf-8') as f:
            for i in f.readlines():
                yield i

    def slide_up(self):
        self.demotext(self.watch_cmd['上滑'])

    def slide_down(self):
        self.demotext(self.watch_cmd['下滑'])

    def slide_left(self):
        self.demotext(self.watch_cmd['左滑'])

    def slide_right(self):
        self.demotext(self.watch_cmd['右滑'])


    def click(self,x_y):

        x, y = hex(int(x_y[0]))[2:], hex(int(x_y[1]))[2:]
        x = '0' * (4 - len(x)) + x
        y = '0' * (4 - len(y)) + y
        s = '7401%s%s0001'% (x, y)
        appcmd="-".join(re.findall(r".{2}", s))
        self.demotext(appcmd)


    def into_music(self):

        self.slide_right()


    def to_homepage(self):
        self.demotext(self.watch_cmd['右键0'])
        self.demotext(self.watch_cmd['右键1'])

    def drink_water(self):
        self.demotext(self.watch_cmd['喝水提醒'])

    def messenger(self):
        self.demotext(self.watch_cmd['消息提醒'])

    def open_music(self):
        self.demotext(self.watch_cmd['打开音乐'])

    def close_music(self):
        self.demotext(self.watch_cmd['关闭音乐'])


    def set_music_1(self,song):


        music_song_singer = '-'.join(re.findall(r'.{2}', (song.encode('utf-8').hex()).ljust(128, '0')))

        music_path = self.watch_cmd['有歌手名称'] + music_song_singer + self.watch_cmd['名称长度'] + music_song_singer + self.watch_cmd[
            '名称长度'] + self.watch_cmd['歌曲状态']

        self.demotext(music_path)


    def set_music_2(self):

        song = self.data.pop(0)

        music_song_singer = '-'.join(re.findall(r'.{2}', (song.encode('utf-8').hex()).ljust(128, '0')))

        music_path = self.watch_cmd['无歌手名称'] + self.watch_cmd['名称长度'] + music_song_singer + self.watch_cmd[
            '名称长度'] + music_song_singer + self.watch_cmd['歌曲状态']

        self.demotext(music_path)



    def get_key_value(self, yamldict):
        for key in yamldict.keys():
            return key, yamldict[key]



    def time_stamp(self, time_str):
        timeArray = time.strptime(time_str, "%Y-%m-%d %H:%M")
        timeStamp = int(time.mktime(timeArray))
        return timeStamp


    def sleep_camera(self, yamlpath, text, s1=1, s2=1):
        time.sleep(s1)
        self.device.camera(yamlpath, text)
        time.sleep(s2)


    def sleep(self):
        time.sleep(1)







    def set_h_unit(self,h_unit):
        if h_unit in ('12',12 ):
            self.demotext('03-11-00-00-00-1e-00-02-1e-00-01')
        elif h_unit in ('24',24 ):
            self.demotext('03-11-00-00-00-1e-00-01-1e-00-01')

    #反转时制
    def reverse_h_unit(self):
        if self.h_unit == '12':
            self.h_unit = '24'
            self.set_h_unit(self.h_unit)
        else:
            self.h_unit = '12'
            self.set_h_unit(self.h_unit)


    def new_excel_row(self):
        self.pic = self.picfield()
        clearStatus()
        self.data = readYaml(self.yamlpath) if readYaml(self.yamlpath) else {}
        self.lang = self.language.pop(0)   # 当前语言
        self.data[self.lang] = []
        writeYaml(self.yamlpath, self.data)

    def save_excel_row(self):
        self.writeExcel(self.pic, self.lang)

    def newExcel(self):
        self.excelname = self.excelname + str(time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())) + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(self.excelname)

    def writeExcel(self, pic, sport):
        '''
        将用例结果写到excel
        '''
        description = self.readtxt()
        wb = load_workbook(self.excelname)
        sheet = wb.worksheets[0]
        nowrow = sheet.max_row + 1
        print('当前行', sheet.max_row)
        sheet.cell(nowrow, 1).value = sport
        piclen = readYaml(self.yamlpath)[sport]
        sheet.row_dimensions[nowrow].height = 250
        for i in range(2, len(piclen) + 2):
            sheet[get_column_letter(i) + str(nowrow)] = next(description)
            if piclen[i - 2] == 's':
                continue
            apic = os.path.join(filePath.FPICTURES_DIR, next(pic))
            if piclen[i - 2] == 'a':
                img = Image.open(apic)
                out = img.resize((int((img.size[0] * 0.35)), int((img.size[1]) * 0.35)))
                out.save(apic)
            sheet.column_dimensions[get_column_letter(i)].width = 40
            img = image.Image(apic)
            sheet.add_image(img, get_column_letter(i) + str(nowrow))
        wb.save(self.excelname)
        wb.close()

    def picfield(self):
        for i in os.listdir(filePath.FPICTURES_DIR):
            yield i







