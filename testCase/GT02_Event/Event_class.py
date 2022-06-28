import os
import time
import openpyxl
import ruamel
import re
import uiautomator2 as u2
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter
from ble.eventremin_cmd import eventCmd
from config.filePath import filePath
from lib.OtherMethod import writeType
from lib.getConfig import getConfig
from lib.yamlMethod import writeYaml, readYaml
from method.OperationDir import clearStatus, delDir
from photo.mainPhoto import mainPhoto

class Event:
    def __init__(self):
        self.demo = u2.connect(getConfig('deviceid', 'app'))
        # self.device = mainPhoto()
        self.PJ_name = 'gt02'
        self.yamlpath = 'event.yaml'
        self.yamlNO = 'eventNO'
        self.event_dataNO_list = []
        self.data = {}
        clearStatus()
        writeYaml(self.yamlpath)
        self.excelname = self.PJ_name+'_event'
        self.newExcel()
        self.watch_cmd = readYaml(filePath.watch_cmd)[self.PJ_name]
        self.event_operate = self.watch_cmd['event_operate']
        self.h_unit = '12'   #时制
        self.demo_apkname = getConfig('demoname', 'app')
        pass

    def demotext(self, text):
        time.sleep(0.5)
        self.demo.app_start(getConfig('demoname', 'app'))
        time.sleep(0.5)
        if self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').exists(1):
            self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').click()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').clear_text()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').set_text(text)
        self.demo.press('back')
        self.demo(text='发送自定义命令').click()

    def readtxt(self):
        with open(filePath.DESCRIPTION, 'r', encoding='utf-8') as f:
            for i in f.readlines():
                yield i

    def slide_up(self):
        self.demotext(self.watch_cmd['上滑'])

    def slide_down(self):
        self.demotext(self.watch_cmd['下滑'])

    def slide_left(self):
        self.demotext(self.watch_cmd['左滑'])

    def slide_right(self):
        self.demotext(self.watch_cmd['右滑'])

    def event_rights(self):
        self.demotext(self.watch_cmd['event_right'])
        self.demotext(self.watch_cmd['event_right'])

    def click(self,x_y):
        x, y = hex(int(x_y[0]))[2:], hex(int(x_y[1]))[2:]
        x = '0' * (4 - len(x)) + x
        y = '0' * (4 - len(y)) + y
        s = '7401%s%s0001'% (x, y)
        appcmd="-".join(re.findall(r".{2}", s))
        self.demotext(appcmd)

    def re_page(self):
        self.slide_right()
        self.click([150,50])


    #设置事件提醒时间到
    def set_time_rem(self,event_rem):
        self.demotext(eventCmd.set_time(event_rem['event_time']))


    def touch_rem(self, event_rem):
        self.set_event(event_rem)
        self.set_time_rem(event_rem)

    def get_key_value(self, yamldict):
        for key in yamldict.keys():
            return key, yamldict[key]

    def time_stamp(self, time_str):
        timeArray = time.strptime(time_str, "%Y-%m-%d %H:%M")
        timeStamp = int(time.mktime(timeArray))
        return timeStamp

    def sleep_camera(self, yamlpath, text, s1=1, s2=1):
        time.sleep(s1)
        self.device.camera(yamlpath, text)
        time.sleep(s2)

    def sleep(self):
        time.sleep(1)

    def time_add(self, time_str, ymdhn, n):
        timeArray = time.strptime(time_str, "%Y/%m/%d/%H/%M")
        timeStamp = int(time.mktime(timeArray))
        if ymdhn == 'y':
            timeStamp = timeStamp + n * 365 * 24 * 60 * 60
        elif ymdhn == 'm':
            timeStamp = timeStamp + n * 30 * 24 * 60 * 60
        elif ymdhn == 'd':
            timeStamp = timeStamp + n * 24 * 60 * 60
        elif ymdhn == 'h':
            timeStamp = timeStamp + n * 60 * 60
        elif ymdhn == 'n':
            timeStamp = timeStamp + n * 60
        timeArray = time.localtime(timeStamp)
        time_str = time.strftime("%Y/%m/%d/%H/%M", timeArray)
        return time_str

    def date_add(self,date_str,ymdhn,n_list):
        if type(n_list) == int:
            date_str = self.time_add(date_str,ymdhn, n_list)
        else:
            for i in range(len(ymdhn)):
                date_str = self.time_add(date_str,ymdhn[i],n_list[i])
        return date_str

    def set_h_unit(self,h_unit):
        if h_unit in ('12',12 ):
            self.demotext('03-11-00-00-00-1e-00-02-1e-00-01')
        elif h_unit in ('24',24 ):
            self.demotext('03-11-00-00-00-1e-00-01-1e-00-01')

    #反转时制
    def reverse_h_unit(self):
        if self.h_unit == '12':
            self.h_unit = '24'
            self.set_h_unit(self.h_unit)
        else:
            self.h_unit = '12'
            self.set_h_unit(self.h_unit)

    def set_event(self,event_rem):
        cmd = eventCmd.event_cmd(event_rem['event_time'],event_rem['event_NO'],event_rem['event_head'],event_rem['event_body'])
        for c in cmd:
            self.demotext(c)
        #判断是否有时制 h_unit 字段
        if 'h_unit' in event_rem.keys():
            self.set_h_unit(event_rem['h_unit'])
            self.h_unit = event_rem['h_unit']

    def set_events(self, event_rem_list):
        event_rem = eval('self.' + event_rem_list[0])
        event_time = event_rem['event_time']
        time_str = self.date_add(event_time, 'd', -1)
        self.demotext(eventCmd.set_time(time_str))
        for i in range(len(event_rem_list)):
            event_rem = eval('self.'+event_rem_list[i])
            self.set_event(event_rem)
        self.re_page()


    def del_all_events(self):
        print('删除所有事件提醒...')
        event_cmd_del_list = eventCmd.event_cmd_del_all()  # 删除所有事件提醒
        for cmd in event_cmd_del_list:
            for c in cmd:
                self.demotext(c)

    def new_excel_row(self):
        self.pic = self.picfield()
        clearStatus()
        self.data = readYaml(self.yamlpath) if readYaml(self.yamlpath) else {}
        self.lang = self.language.pop(0)   # 当前语言
        self.data[self.lang] = []
        writeYaml(self.yamlpath, self.data)

    def save_excel_row(self):
        self.writeExcel(self.pic, self.lang)

    def newExcel(self):
        self.excelname = self.excelname + str(time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())) + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(self.excelname)

    def writeExcel(self, pic, sport):
        '''
        将用例结果写到excel
        '''
        description = self.readtxt()
        wb = load_workbook(self.excelname)
        sheet = wb.worksheets[0]
        nowrow = sheet.max_row + 1
        print('当前行', sheet.max_row)
        sheet.cell(nowrow, 1).value = sport
        piclen = readYaml(self.yamlpath)[sport]
        sheet.row_dimensions[nowrow].height = 250
        for i in range(2, len(piclen) + 2):
            sheet[get_column_letter(i) + str(nowrow)] = next(description)
            if piclen[i - 2] == 's':
                continue
            apic = os.path.join(filePath.FPICTURES_DIR, next(pic))
            if piclen[i - 2] == 'a':
                img = Image.open(apic)
                out = img.resize((int((img.size[0] * 0.35)), int((img.size[1]) * 0.35)))
                out.save(apic)
            sheet.column_dimensions[get_column_letter(i)].width = 40
            img = image.Image(apic)
            sheet.add_image(img, get_column_letter(i) + str(nowrow))
        wb.save(self.excelname)
        wb.close()

    def picfield(self):
        for i in os.listdir(filePath.FPICTURES_DIR):
            yield i



