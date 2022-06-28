import os
import time
import openpyxl
import ruamel
import re
import uiautomator2 as u2
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter
from ble.eventremin_cmd import eventCmd
from config.filePath import filePath
from lib.OtherMethod import writeType
from lib.getConfig import getConfig
from lib.yamlMethod import writeYaml, readYaml
from method.OperationDir import clearStatus, delDir
from photo.mainPhoto import mainPhoto

'''
    说明：
        1. 编码：longt
        2. 开始时间：2022-6-14
        3. 最近更新：2020-6-14
        4. 简介：运动记录模块基本类
'''

class SportExcel:


    def __init__(self):


        self.demo = u2.connect(getConfig('deviceid', 'app'))
        self.device = mainPhoto()
        self.PJ_name = 'ID_W03'   # 项目名称
        self.moudle = 'sport_record'    # 模块名称
        self.cmd = 'all'
        self.yamlpath = 'sport_record_carema.yaml'
        self.yamlNO = 'sport_record_NO'
        self.music_data_NO_list = []
        self.data = {}
        clearStatus()
        writeYaml(self.yamlpath)
        self.excelname = self.PJ_name + self.moudle
        self.newExcel()
        self.watch_cmd = readYaml(filePath.all_cmd)[self.cmd]
        self.h_unit = '12'   #时制
        self.demo_apkname = getConfig('demoname', 'app')
        pass


    def demotext(self, text):

        time.sleep(0.5)
        self.demo.app_start(getConfig('demoname', 'app'))
        time.sleep(0.5)
        if self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').exists(1):
            self.demo(resourceId='test.com.ido:id/home_btn_app_control_device').click()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').clear_text()
        self.demo(resourceId='test.com.ido:id/custom_cmd_data_et').set_text(text)
        time.sleep(0.2)
        self.demo.press('back')
        self.demo(text='发送自定义命令').click()


    def readtxt(self):
        with open(filePath.DESCRIPTION, 'r', encoding='utf-8') as f:
            for i in f.readlines():
                yield i

    def slide_up(self):
        self.demotext(self.watch_cmd['上滑'])

    def slide_up_2(self):
        self.demotext(self.watch_cmd['整屏上滑'])


    def slide_down(self):
        self.demotext(self.watch_cmd['下滑'])

    def slide_left(self):
        self.demotext(self.watch_cmd['左滑'])

    def slide_right(self):
        self.demotext(self.watch_cmd['右滑'])

    def to_homepage(self):
        self.demotext(self.watch_cmd['右键0'])
        self.demotext(self.watch_cmd['右键1'])

    def set_dis(self, type):
        if str(type[0])== '1':
            self.demotext(self.watch_cmd['公制'])
        elif str(type[0]) == '0':
            self.demotext(self.watch_cmd['英制'])



    def click(self,x_y):

        x, y = hex(int(x_y[0]))[2:], hex(int(x_y[1]))[2:]
        x = '0' * (4 - len(x)) + x
        y = '0' * (4 - len(y)) + y
        s = '7401%s%s0001'% (x, y)
        appcmd="-".join(re.findall(r".{2}", s))
        self.demotext(appcmd)



    '''进入运动1'''
    def into_sport_1(self):
        self.to_homepage()
        self.click([150,160])


    def close_sport(self):
        self.slide_right()
        self.click([150,180])
        self.slide_right()
        self.slide_right()
        self.slide_right()

    def close_sport_2(self):
        self.slide_right()
        self.click([150,180])
        self.slide_right()
        self.slide_right()

    def ringt_3(self):
        for _ in range(3):
            self.slide_right()




    '''进入运动列表'''
    def into_sport_list(self):
        self.to_homepage()
        self.slide_up()
        self.click([150,250])

    def into_sport_list_2(self):
        self.slide_up()
        self.click([150,250])



    '''设置运动中数据'''
    def set_data_daily(self,type_number):
        data_type = str(type_number[0])
        data_1 = str(hex(int(type_number[1])))[2:].rjust(8, '0')
        data_2 = data_1[-2:] + '-' + data_1[-4:-2] + '-' + data_1[-6:-4] + '-' + data_1[-8:-6]
        dict = {'step': '74-08-10-' + data_2, 'heart': '74-08-11-' + data_2, 'pace': '74-08-12-' + data_2,
                'distance': '74-08-13-' + data_2, 'kcal': '74-08-14-' + data_2, 'age_pace': '74-08-15-' + data_2}
        if data_type in dict.keys():
            self.demotext(dict[data_type])
        else:
            print('error')


    '''设置运动类型'''
    def set_sport(self,type):

        stop_dict_1 = {'户外跑步': '30', '室内跑步': '31', '户外步行': '34', '室内步行': '35', '户外骑行': '32', '室内骑行': '33', '泳池游泳': '36',
                     '开放水域游泳': '37', '徒步': '04', '瑜伽': '12', '划船机': '39', '椭圆机': '38', '板球': '4B', '健身': '08',
                     '高强度间歇训练': '3A', '功能性力量训练': '65', '核心训练': '66', '舞蹈': '1D', '踏步机': '67', '整理放松': '68',
                     '传统力量训练': '6E', '有氧健身操': '11', '普拉提': '20', '仰卧起坐': '0D', '平板支撑': '25', '开合跳': '72', '引体向上': '70',
                     '俯卧撑': '0E', '深蹲': '73', '高抬腿': '74', '哑铃': '0F', '杠铃': '76', '拳击': '75', '武术': '77', '太极': '78',
                     '跆拳道': '79', '空手道': '7A', '自由搏击': '7B', '击剑': '7C', '射箭': '7D', '体操': '7E', '单杠': '7F', '双杠': '80',
                     '漫步机': '81', '登山机': '82',
                     '保龄球': '83', '网球': '18', '乒乓球': '14', '高尔夫球': '19', '足球': '16', '篮球': '15', '台球': '84',
                     '羽毛球': '07', '曲棍球': '85', '橄榄球': '86', '壁球': '87', '垒球': '88', '手球': '89', '毽球': '8A', '棒球': '1A',
                     '沙滩足球': '8B', '藤球': '8C', '躲避球': '8D',
                     '广场舞': '24', '街舞': '98', '芭蕾': '99', '社交舞': '9A', '登山': '06', '跳绳': '13', '飞盘': '9B', '飞镖': '9C',
                     '骑马': '9D', '爬楼': '9E', '放风筝': '9F', '钓鱼': 'A0',
                     '雪橇': 'A1', '雪车': 'A2', '滑冰': '1B', '单板滑雪': 'A3', '高山滑雪': 'A5', '越野滑雪': 'A6', '雪上运动': 'A4',
                     '冰壶': 'A7', '冰球': 'A8', '冬季两项': 'A9',
                     '冲浪': 'AA', '帆船': 'AB', '帆板': 'AC', '皮艇': 'AD', '划艇': 'AF', '赛艇': 'B0', '摩托艇': 'AE', '龙舟': 'B1',
                     '水球': 'B2', '漂流': 'B3',
                     '滑板': 'B4', '攀岩': 'B5', '蹦极': 'B6', '跑酷': 'B7', 'BMX': 'B8', }

        sport_dict_2 = {'户外骑行': '1', '户外行走': '2', '户外跑步': '3', '室内跑步': '4', '登山': '5', '徒步': '6', '瑜伽': '8', '板球': '9',
             '室内骑行': 'a', '椭圆机': 'b', '篮球': 'd', '足球': 'e', '乒乓球': 'f', '羽毛球': '10', '走路': '11', '划船机': '12',
             '自由训练': '13', '力量训练': '16', '曲棍球': '17', '网球': '18', '': '19', '骑马': '1b', '太极': '1c', '毽球': '1d',
             '拳击': '1e', '越野跑': '1f', '体操': '20', '冰上曲棍球': '21', '跆拳道': '23', '舞蹈': '25', '交叉训练': '27', '普拉提': '28',
             '交叉配合': '29', '跳绳': '2c', '射箭': '2d', '街舞': '31', '自由搏击': '32', '芭蕾': '33', '武术': '35', '爬楼': '36',
             '手球': '37', '深蹲': '70', '高抬腿': '71', '杠铃': '72', '功能性力量训练': '73', '广场舞': '74', '社交舞': '75',
             '沙滩足球': '76', '台球': '77', '滕球': '78', '躲避球': '79', '赛艇': '7a', '摩托艇': '7b', '龙舟': '7c', '冲浪': '7d',
             '棒球': '38', '保龄球': '39', '壁球': '3a', '冰壶': '3b', '单板滑雪': '3d', '美式足球': '3f', '钓鱼': '41', '飞盘运动': '42',
             '高山滑雪': '44', '雪上运动': '45', '核心训练': '48', '滑冰': '49', '团体操': '4c', '长曲棍球': '4e', '击剑': '51',
             '垒球': '52', '单杠': '53', '双杠': '54', '太空漫步机': '55', '飞镖': '57', '仰卧起坐': '59', '空手道': '5c', '整理放松': '5d',
             '橄榄球': '5e', '高二夫': '5f', '定向越野': '61', '攀岩': '62', '滑板': '63', '卡巴迪': '64', '波比跳': '65', '踏步机': '66',
             '卷腹': '67', '有氧健身操': '68', '蹦床': '69', '俯卧撑': '6a', '引体向上': '6b', '平板向上': '6c', '哑铃': '6d', '登山机': '6e',
             '开合跳': '6f', '帆板': '7e', '帆船': '7f', '皮艇': '80', '划船': '81', '漂流': '82', '越野滑雪': '83', '冬季两项': '84',
             '雪车': '85', '雪橇': '86', '赛车': '87', '跳伞': '88', '高空弹跳': '89', '跑酷': '8a', 'BMX': '8b'}

        sport_frst = '33-da-ad-da-ad-01-7d-00-41-00-13-00-00-02-03-6e-03-04-'
        sport_tend = sport_dict_2[str(type[0])].rjust(2,'0')
        sport_lt = '-11-01-16-0e-17-0f-10-0b-08-09-1c-1d-1e-0d-05-1f-20-21-23-12-55-06-18-25-5c-5d-27-28-2c-2d-31-32-33' \
                     '-35-36-37-38-39-3a-3b-3d-3f-41-42-5e-5f-44-45-48-49-4c-4e-51-52-53-54-60-57-59-19-02-0a-61-62-63-' \
                     '64-65-66-67-68-69-6a-6b-6c-6d-6e-6f-70-71-72-73-74-75-76-77-78-79-7a-7b-7c-7d-7e-7f-80-81-82-83-84' \
                     '-85-86-87-88-89-8a-8b-8c-8d-a1-6a'
        sport_type = sport_frst + sport_tend + sport_lt
        self.demotext(str(sport_type))


    def set_time(self,date):
        '''
        设置时间,格式1: 2021/12/27/09/00'
                格式2: 2021/12/27
        '''
        if len(str(date)) > 11:
            years, months, days, hour, minute,= str(date).split('/')
            cmd = "03-01-" + self.ten_to_hex(years) + "-" + self.ten_to_hex(months) + "-" + self.ten_to_hex(days) + "-" + self.ten_to_hex(hour) + "-" + self.ten_to_hex(minute) + "-38-ff"
        else:
            years, months, days, = str(date).split('/')
            cmd = "03-01-" + self.ten_to_hex(years) + "-" + self.ten_to_hex(months) + "-" + self.ten_to_hex(days) + "-09-00-38-ff"
        self.demotext(str(cmd))


    def set_time_exe(self,date_var_str):

        eval('self.set_time("'+ str(date_var_str)+ '")' )



    def ten_to_hex(self, num, several=0):
        """
        num为待转换为16进制数据，
        text为excel单元格中读取到的命令
        serveral 为uint32_t value
        """
        if several == 0:
            several = len(str(num))
            if len(str(num)) % 2 == 1:
                num = "0" + str(num)
                several = len(str(num))
        trans_code = '{:0>{several}x}'.format(int(num), several=several)  # 小端序  several设置位数
        instruct = ''
        for i in range(len(trans_code), 0, -2):
            instruct = "-".join([instruct, trans_code[i - 2:i]])
        return instruct[1:]





    def get_key_value(self, yamldict):
        for key in yamldict.keys():
            return key, yamldict[key]



    def time_stamp(self, time_str):
        timeArray = time.strptime(time_str, "%Y-%m-%d %H:%M")
        timeStamp = int(time.mktime(timeArray))
        return timeStamp


    def sleep_camera(self, yamlpath, text, s1=1, s2=1):
        time.sleep(s1)
        self.device.camera(yamlpath, text)
        time.sleep(s2)


    def sleep(self):
        time.sleep(1.5)



    def set_h_unit(self,h_unit):
        if h_unit in ('12',12 ):
            self.demotext('03-11-00-00-00-1e-00-02-1e-00-01')
        elif h_unit in ('24',24 ):
            self.demotext('03-11-00-00-00-1e-00-01-1e-00-01')

    #反转时制
    def reverse_h_unit(self):
        if self.h_unit == '12':
            self.h_unit = '24'
            self.set_h_unit(self.h_unit)
        else:
            self.h_unit = '12'
            self.set_h_unit(self.h_unit)


    def new_excel_row(self):
        self.pic = self.picfield()
        clearStatus()
        self.data = readYaml(self.yamlpath) if readYaml(self.yamlpath) else {}
        self.lang = self.language.pop(0)   # 当前语言
        self.data[self.lang] = []
        writeYaml(self.yamlpath, self.data)


    def save_excel_row(self):
        self.writeExcel(self.pic, self.lang)



    def newExcel(self):
        self.excelname = self.excelname + str(time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())) + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(self.excelname)



    def writeExcel(self, pic, sport):
        '''
        将用例结果写到excel
        '''
        description = self.readtxt()
        wb = load_workbook(self.excelname)
        sheet = wb.worksheets[0]
        nowrow = sheet.max_row + 1
        print('当前行', sheet.max_row)
        sheet.cell(nowrow, 1).value = sport
        piclen = readYaml(self.yamlpath)[sport]
        sheet.row_dimensions[nowrow].height = 250
        for i in range(2, len(piclen) + 2):
            sheet[get_column_letter(i) + str(nowrow)] = next(description)
            if piclen[i - 2] == 's':
                continue
            apic = os.path.join(filePath.FPICTURES_DIR, next(pic))
            if piclen[i - 2] == 'a':
                img = Image.open(apic)
                out = img.resize((int((img.size[0] * 0.35)), int((img.size[1]) * 0.35)))
                out.save(apic)
            sheet.column_dimensions[get_column_letter(i)].width = 40
            img = image.Image(apic)
            sheet.add_image(img, get_column_letter(i) + str(nowrow))
        wb.save(self.excelname)
        wb.close()

    def picfield(self):
        for i in os.listdir(filePath.FPICTURES_DIR):
            yield i





